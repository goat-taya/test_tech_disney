from openpyxl import Workbook, load_workbook

from translator.pipeline import translate_file


def test_pipeline_dispatches_by_extension(tmp_path):
    source = tmp_path / "products.xlsx"
    output_dir = tmp_path / "output"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["sku", "description"])
    sheet.append(["A-001", "Texte francais"])
    workbook.save(source)

    translated_path = translate_file(source, output_dir)

    assert translated_path == output_dir / "products_en.xlsx"
    rows = list(load_workbook(translated_path).active.iter_rows(values_only=True))
    assert rows[1][2] == "[EN] Texte francais"
