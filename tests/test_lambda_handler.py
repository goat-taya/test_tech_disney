import json

from openpyxl import Workbook, load_workbook

from translator.lambda_handler import handler


def test_lambda_handler_processes_local_file_event(tmp_path):
    input_dir = tmp_path / "input"
    output_dir = tmp_path / "output"
    input_dir.mkdir()
    source = input_dir / "products.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["sku", "description"])
    sheet.append(["A-001", "Texte commercial"])
    workbook.save(source)

    response = handler(
        {
            "storage": "local",
            "input_path": str(source),
            "output_dir": str(output_dir),
        },
        None,
    )

    assert response["statusCode"] == 200
    body = json.loads(response["body"])
    translated_path = output_dir / "products_en.xlsx"
    assert body["translated_file"] == str(translated_path)
    rows = list(load_workbook(translated_path).active.iter_rows(values_only=True))
    assert rows[1][2] == "[EN] Texte commercial"
