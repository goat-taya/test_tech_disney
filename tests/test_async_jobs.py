import json

from openpyxl import Workbook, load_workbook

from translator.async_handler import handler as async_handler
from translator.jobs import JobStore
from translator.worker import process_available_jobs, process_next_job


def test_async_handler_enqueues_job_and_worker_processes_it(tmp_path):
    source = tmp_path / "input" / "products.xlsx"
    output_dir = tmp_path / "output"
    job_dir = tmp_path / "jobs"
    source.parent.mkdir()

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["sku", "description"])
    sheet.append(["A-001", "Texte scalable"])
    workbook.save(source)

    response = async_handler(
        {
            "storage": "local",
            "input_path": str(source),
            "output_dir": str(output_dir),
            "job_dir": str(job_dir),
        },
        None,
    )

    assert response["statusCode"] == 202
    body = json.loads(response["body"])
    assert body["status"] == "queued"

    store = JobStore(job_dir)
    queued_job = store.get(body["job_id"])
    assert queued_job["status"] == "queued"

    processed = process_next_job(store)

    assert processed["job_id"] == body["job_id"]
    assert processed["status"] == "succeeded"
    rows = list(load_workbook(output_dir / "products_en.xlsx").active.iter_rows(values_only=True))
    assert rows[1][2] == "[EN] Texte scalable"


def test_worker_marks_job_failed_when_translation_errors(tmp_path):
    source = tmp_path / "missing.xlsx"
    store = JobStore(tmp_path / "jobs")
    job = store.enqueue(source, tmp_path / "output")

    processed = process_next_job(store)

    assert processed["job_id"] == job["job_id"]
    assert processed["status"] == "failed"
    assert "error" in processed


def test_async_handler_can_return_job_status(tmp_path):
    source = tmp_path / "input" / "products.xlsx"
    output_dir = tmp_path / "output"
    job_dir = tmp_path / "jobs"
    source.parent.mkdir()

    workbook = Workbook()
    workbook.active.append(["sku", "description"])
    workbook.save(source)

    response = async_handler(
        {
            "storage": "local",
            "input_path": str(source),
            "output_dir": str(output_dir),
            "job_dir": str(job_dir),
        },
        None,
    )
    job_id = json.loads(response["body"])["job_id"]

    status_response = async_handler(
        {
            "action": "status",
            "storage": "local",
            "job_id": job_id,
            "job_dir": str(job_dir),
        },
        None,
    )

    assert status_response["statusCode"] == 200
    body = json.loads(status_response["body"])
    assert body["job_id"] == job_id
    assert body["status"] == "queued"


def test_worker_requeues_retryable_failures_before_terminal_failure(tmp_path, monkeypatch):
    source = tmp_path / "input" / "products.xlsx"
    output_dir = tmp_path / "output"
    store = JobStore(tmp_path / "jobs")
    source.parent.mkdir()

    workbook = Workbook()
    workbook.active.append(["sku", "description"])
    workbook.active.append(["A-001", "Texte retry"])
    workbook.save(source)

    job = store.enqueue(source, output_dir, max_attempts=2)
    calls = []

    def flaky_translate_file(input_path, output_path, service=None):
        calls.append(input_path)
        if len(calls) == 1:
            raise RuntimeError("temporary outage")
        from translator.pipeline import translate_file

        return translate_file(input_path, output_path, service=service)

    monkeypatch.setattr("translator.worker.translate_file", flaky_translate_file)

    first = process_next_job(store)
    assert first["job_id"] == job["job_id"]
    assert first["status"] == "retry_scheduled"
    assert first["attempts"] == 1
    assert first["last_error"] == "temporary outage"
    assert "retry_available_at" in first

    second = process_next_job(store)
    assert second["status"] == "succeeded"
    assert second["attempts"] == 2


def test_worker_does_not_process_retries_before_they_are_due(tmp_path, monkeypatch):
    source = tmp_path / "input" / "products.xlsx"
    output_dir = tmp_path / "output"
    store = JobStore(tmp_path / "jobs")
    source.parent.mkdir()

    workbook = Workbook()
    workbook.active.append(["sku", "description"])
    workbook.active.append(["A-001", "Texte delayed retry"])
    workbook.save(source)

    store.enqueue(source, output_dir, max_attempts=2, retry_delay_seconds=60)

    def failing_translate_file(input_path, output_path, service=None):
        raise RuntimeError("temporary outage")

    monkeypatch.setattr("translator.worker.translate_file", failing_translate_file)

    first = process_next_job(store)
    second = process_next_job(store)

    assert first["status"] == "retry_scheduled"
    assert second is None


def test_worker_processes_multiple_queued_jobs(tmp_path):
    output_dir = tmp_path / "output"
    store = JobStore(tmp_path / "jobs")

    for index in range(2):
        source = tmp_path / "input" / f"products_{index}.xlsx"
        source.parent.mkdir(exist_ok=True)
        workbook = Workbook()
        workbook.active.append(["sku", "description"])
        workbook.active.append([f"A-00{index}", f"Texte {index}"])
        workbook.save(source)
        store.enqueue(source, output_dir)

    processed = process_available_jobs(store, max_jobs=10)

    assert [job["status"] for job in processed] == ["succeeded", "succeeded"]
