from pathlib import Path

from openpyxl import load_workbook

from translator.service import TranslationService


def translate_excel(
    input_path: str | Path,
    output_path: str | Path,
    service: TranslationService | None = None,
) -> Path:
    input_path = Path(input_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    service = service or TranslationService()

    workbook = load_workbook(input_path)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    if "description" not in headers:
        raise ValueError("Excel file must contain a 'description' column")

    description_col = headers.index("description") + 1
    translated_col = description_col + 1

    if translated_col > len(headers) or headers[translated_col - 1] != "description_en":
        sheet.insert_cols(translated_col)
        sheet.cell(row=1, column=translated_col, value="description_en")

    row_values = []
    row_numbers = []
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row, column=description_col).value
        if value is not None and str(value).strip():
            row_values.append(str(value))
            row_numbers.append(row)

    translations = service.translate_many(row_values)
    for row, translated_text in zip(row_numbers, translations, strict=True):
        sheet.cell(row=row, column=translated_col, value=translated_text)

    workbook.save(output_path)
    return output_path
