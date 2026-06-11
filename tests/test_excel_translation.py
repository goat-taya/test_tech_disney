from openpyxl import Workbook, load_workbook

from translator.excel import translate_excel
from translator.service import TranslationService


def test_excel_adds_description_en_next_to_description_and_preserves_other_columns(tmp_path):
    source = tmp_path / "products.xlsx"
    output = tmp_path / "products_en.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["sku", "description", "category"])
    sheet.append(["A-001", "Figurine edition limitee", "jouets"])
    sheet.append(["A-002", "Mug pour les equipes internes", "accessoires"])
    workbook.save(source)

    translate_excel(source, output)

    translated = load_workbook(output)
    rows = list(translated.active.iter_rows(values_only=True))

    assert rows[0] == ("sku", "description", "description_en", "category")
    assert rows[1] == ("A-001", "Figurine edition limitee", "[EN] Figurine edition limitee", "jouets")
    assert rows[2] == (
        "A-002",
        "Mug pour les equipes internes",
        "[EN] Mug pour les equipes internes",
        "accessoires",
    )


def test_excel_uses_translation_service_cache_for_duplicate_descriptions(tmp_path):
    source = tmp_path / "products.xlsx"
    output = tmp_path / "products_en.xlsx"
    calls = []

    def backend(text, source="fr", target="en"):
        calls.append(text)
        return f"[EN] {text}"

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["sku", "description"])
    sheet.append(["A-001", "Description partagee"])
    sheet.append(["A-002", "Description partagee"])
    workbook.save(source)

    translate_excel(source, output, service=TranslationService(backend=backend))

    assert calls == ["Description partagee"]
